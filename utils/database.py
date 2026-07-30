# ============================================================
# database.py — SQLite Database Manager
# Full schema matching the AstraQuote Lead Engine pipeline
# ============================================================

import sqlite3
import os
import csv
import json
import shutil
from typing import Dict, List, Optional, Any
from datetime import datetime

from utils.logger import get_logger

logger = get_logger("Database")

# Default path — will be overridden by init_db()
_db_path = "data/leads.db"

SCHEMA = """
CREATE TABLE IF NOT EXISTS leads (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    company_name    TEXT NOT NULL,
    canton          TEXT,
    city            TEXT,
    niche           TEXT,
    phone           TEXT UNIQUE,
    email           TEXT,
    website         TEXT,
    address         TEXT,
    postal_code     TEXT,
    noga_code       TEXT,

    -- Company Info
    legal_form      TEXT,
    founded_year    INTEGER,
    employee_count  INTEGER,
    zefix_uid       TEXT,
    zefix_status    TEXT,
    business_description TEXT,

    -- Company Size Estimate (see utils/company_size.py)
    size_band       TEXT,
    employees_estimate TEXT,
    size_signals    TEXT,
    officer_count   INTEGER,
    team_headcount_hint INTEGER,

    -- Digital Presence
    has_website     BOOLEAN DEFAULT 0,
    has_quote_form  BOOLEAN DEFAULT 0,
    has_instagram   TEXT,
    has_facebook    TEXT,
    has_linkedin    TEXT,
    google_rating   REAL,
    google_reviews  INTEGER,
    digital_maturity INTEGER DEFAULT 0,

    -- Research Analysis
    pain_points     TEXT,
    pitch_angle     TEXT,
    custom_opening  TEXT,
    urgency_score   INTEGER DEFAULT 0,
    estimated_quotes INTEGER,
    research_complete BOOLEAN DEFAULT 0,

    -- Contact Enrichment
    decision_maker  TEXT,
    decision_title  TEXT,
    decision_maker_linkedin TEXT,
    contact_score   INTEGER DEFAULT 0,
    is_mobile       BOOLEAN DEFAULT 0,

    -- AstraQuote ICP Fit
    fit_score       INTEGER DEFAULT 0,
    fit_score_breakdown TEXT,

    -- Pipeline
    source          TEXT,
    status          TEXT DEFAULT 'discovered',
    layer_reached   INTEGER DEFAULT 1,
    raw_snippet     TEXT,
    elimination_reasons TEXT,
    run_id          TEXT,

    -- Timestamps
    discovered_at   TEXT DEFAULT (datetime('now')),
    researched_at   TEXT,
    enriched_at     TEXT,
    created_at      TEXT DEFAULT (datetime('now')),
    updated_at      TEXT DEFAULT (datetime('now'))
);

CREATE INDEX IF NOT EXISTS idx_leads_status ON leads(status);
CREATE INDEX IF NOT EXISTS idx_leads_canton ON leads(canton);
CREATE INDEX IF NOT EXISTS idx_leads_phone ON leads(phone);

CREATE TABLE IF NOT EXISTS blacklist (
    phone           TEXT PRIMARY KEY,
    company_name    TEXT,
    name_key        TEXT,
    added_at        TEXT DEFAULT (datetime('now'))
);
"""
# idx_blacklist_name_key is created in _migrate_schema(), AFTER the guarded
# ALTER TABLE — not here. On a database that predates name_key, an index on
# that column would fail with "no such column" if it ran as part of this
# static script, since CREATE TABLE IF NOT EXISTS is a no-op on an existing
# table and never adds the column itself (that's what _migrate_table does).
# That exact failure aborted the surrounding executescript() call, which
# silently skipped _migrate_schema() and conn.commit() entirely afterward.


def get_connection():
    """Get a database connection."""
    db_dir = os.path.dirname(_db_path)
    if db_dir and not os.path.exists(db_dir):
        os.makedirs(db_dir, exist_ok=True)
    conn = sqlite3.connect(_db_path)
    conn.row_factory = sqlite3.Row
    return conn


# Columns added after the initial release — kept here so existing databases get
# migrated in place instead of silently missing columns new code expects.
_MIGRATION_COLUMNS = {
    "decision_maker_linkedin": "TEXT",
    "business_description": "TEXT",
    "zefix_status": "TEXT",
    "fit_score": "INTEGER",
    "fit_score_breakdown": "TEXT",
    "research_complete": "BOOLEAN DEFAULT 0",
    "size_band": "TEXT",
    "employees_estimate": "TEXT",
    "size_signals": "TEXT",
    "officer_count": "INTEGER",
    "team_headcount_hint": "INTEGER",
    "run_id": "TEXT",
    "noga_code": "TEXT",
}

# name_key = normalized company name (see layers/layer1_discovery.normalize_name),
# stored alongside phone so a blacklisted company can be recognized again even
# under a DIFFERENT phone number (e.g. the owner's mobile vs the shop landline) —
# closes the gap where phone-only blacklist checks let the same company back in
# as an apparently-new lead. See is_blacklisted_by_name().
_BLACKLIST_MIGRATION_COLUMNS = {
    "name_key": "TEXT",
}


def _migrate_table(conn: sqlite3.Connection, table: str, columns: Dict[str, str]) -> None:
    """Additively bring an existing table up to date. CREATE TABLE IF NOT EXISTS
    only helps on a brand-new database — on an existing one it's a no-op, so
    columns introduced later need an explicit guarded ALTER TABLE."""
    existing_cols = {row[1] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    for col, col_type in columns.items():
        if col not in existing_cols:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {col_type}")
            logger.info(f"Migrated {table} table: added column '{col}'")


def _migrate_schema(conn: sqlite3.Connection) -> None:
    _migrate_table(conn, "leads", _MIGRATION_COLUMNS)
    _migrate_table(conn, "blacklist", _BLACKLIST_MIGRATION_COLUMNS)
    # Only safe to create AFTER the ALTER TABLE above guarantees the column exists.
    conn.execute("CREATE INDEX IF NOT EXISTS idx_blacklist_name_key ON blacklist(name_key)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_leads_run_id ON leads(run_id)")


def init_db(path: str = None) -> None:
    """Initialize the database schema."""
    global _db_path
    if path:
        _db_path = path

    db_dir = os.path.dirname(_db_path)
    if db_dir and not os.path.exists(db_dir):
        os.makedirs(db_dir, exist_ok=True)

    conn = sqlite3.connect(_db_path)
    try:
        conn.executescript(SCHEMA)
        _migrate_schema(conn)
        conn.commit()
        logger.info(f"Database initialized at {_db_path}")
    except Exception as e:
        logger.error(f"Error initializing DB: {e}")
    finally:
        conn.close()


def insert_lead(lead_dict: Dict[str, Any]) -> None:
    """Insert or update a lead by phone number (upsert)."""
    phone = lead_dict.get("phone")
    if not phone:
        return

    # Map common aliases
    field_map = {
        "company_name": "company_name",
        "name": "company_name",
    }
    mapped = {}
    for k, v in lead_dict.items():
        key = field_map.get(k, k)
        # Skip non-column keys (transient in-memory fields the pipeline attaches
        # to a lead but that aren't persisted columns).
        if key in ("elimination_reasons", "needs_size_check", "disqualified",
                    "disqualification_reason", "research", "_dm_candidates",
                    "zefix_legal_form", "mentions_team", "web_dm_candidate_count"):
            continue
        mapped[key] = v

    # Ensure company_name exists
    if "company_name" not in mapped:
        mapped["company_name"] = "Unknown"

    mapped["updated_at"] = datetime.now().isoformat()

    cols = ", ".join(mapped.keys())
    placeholders = ", ".join("?" for _ in mapped)
    updates = ", ".join(f"{k} = excluded.{k}" for k in mapped.keys() if k != "phone")

    sql = f"""
    INSERT INTO leads ({cols})
    VALUES ({placeholders})
    ON CONFLICT(phone) DO UPDATE SET
    {updates}
    """

    conn = get_connection()
    try:
        conn.execute(sql, tuple(mapped.values()))
        conn.commit()
    except Exception as e:
        logger.error(f"Error inserting lead {phone}: {e}")
    finally:
        conn.close()


def update_lead(phone: str, updates: Dict[str, Any]) -> None:
    """Update specific fields for a lead identified by phone."""
    if not phone or not updates:
        return

    updates["updated_at"] = datetime.now().isoformat()
    set_clause = ", ".join(f"{k} = ?" for k in updates.keys())
    sql = f"UPDATE leads SET {set_clause} WHERE phone = ?"

    conn = get_connection()
    try:
        conn.execute(sql, list(updates.values()) + [phone])
        conn.commit()
    except Exception as e:
        logger.error(f"Error updating lead {phone}: {e}")
    finally:
        conn.close()


def get_leads(status: Optional[str] = None, layer: Optional[int] = None,
              canton: Optional[str] = None) -> List[Dict[str, Any]]:
    """Get leads with optional filtering, strictly within the last 6 months."""
    conn = get_connection()
    try:
        # Filter for leads discovered within the last 6 months
        query = "SELECT * FROM leads WHERE discovered_at >= date('now', '-6 months')"
        params: list = []
        if status:
            query += " AND status = ?"
            params.append(status)
        if layer is not None:
            query += " AND layer_reached >= ?"
            params.append(layer)
        if canton:
            query += " AND canton = ?"
            params.append(canton)

        cursor = conn.execute(query, params)
        return [dict(row) for row in cursor.fetchall()]
    finally:
        conn.close()


def get_lead_count(status: Optional[str] = None, min_layer: Optional[int] = None) -> int:
    """Count leads matching criteria, strictly within the last 6 months."""
    conn = get_connection()
    try:
        # Filter for leads discovered within the last 6 months
        query = "SELECT COUNT(*) FROM leads WHERE discovered_at >= date('now', '-6 months')"
        params: list = []
        if status:
            query += " AND status = ?"
            params.append(status)
        if min_layer is not None:
            query += " AND layer_reached >= ?"
            params.append(min_layer)
        return conn.execute(query, params).fetchone()[0]
    finally:
        conn.close()


def get_dashboard_leads() -> List[Dict[str, Any]]:
    """Get all leads for the dashboard."""
    return get_leads()


def get_pending_research() -> List[Dict[str, Any]]:
    """Get leads that passed filter but haven't been researched."""
    return get_leads(status="qualified")


def mark_rejected(phone: str, reason: str) -> None:
    """Mark a lead as rejected."""
    update_lead(phone, {"status": "rejected", "elimination_reasons": reason})


def add_to_blacklist(phone: str, company_name: str, name_key: str = "") -> None:
    """Add a lead to the blacklist so they are never contacted again — by phone
    AND, when the caller supplies it, by normalized company name (`name_key`),
    so the SAME company showing up under a different phone number in a future
    run (owner's mobile vs the shop's landline is the common case) is still
    recognized rather than treated as a fresh lead. `name_key` is computed by
    the caller (layers/layer1_discovery.normalize_name) rather than here, to
    keep this module free of a dependency on the layers package."""
    if not phone:
        return
    try:
        with get_connection() as conn:
            conn.execute(
                "INSERT OR IGNORE INTO blacklist (phone, company_name, name_key) VALUES (?, ?, ?)",
                (phone, company_name, name_key or None)
            )
            conn.commit()
    except Exception as e:
        logger.error(f"Error adding to blacklist: {e}")


def is_blacklisted(phone: str) -> bool:
    """Check if a phone number is blacklisted."""
    if not phone:
        return False
    try:
        with get_connection() as conn:
            cur = conn.execute("SELECT 1 FROM blacklist WHERE phone = ?", (phone,))
            return cur.fetchone() is not None
    except Exception as e:
        logger.error(f"Error checking blacklist: {e}")
        return False


def is_blacklisted_by_name(name_key: str) -> bool:
    """Check if a normalized company name matches a previously-finalized lead —
    catches the same company reappearing under a different phone number, which
    a phone-only blacklist check would miss entirely."""
    if not name_key:
        return False
    try:
        with get_connection() as conn:
            cur = conn.execute("SELECT 1 FROM blacklist WHERE name_key = ?", (name_key,))
            return cur.fetchone() is not None
    except Exception as e:
        logger.error(f"Error checking name blacklist: {e}")
        return False


def get_stats() -> Dict[str, Any]:
    """Get pipeline statistics."""
    conn = get_connection()
    try:
        total = conn.execute("SELECT COUNT(*) FROM leads").fetchone()[0]

        status_counts = {}
        for row in conn.execute("SELECT status, COUNT(*) as cnt FROM leads GROUP BY status").fetchall():
            status_counts[row["status"]] = row["cnt"]

        canton_counts = {}
        for row in conn.execute("SELECT canton, COUNT(*) as cnt FROM leads WHERE canton IS NOT NULL AND canton != '' GROUP BY canton").fetchall():
            canton_counts[row["canton"]] = row["cnt"]

        niche_counts = {}
        for row in conn.execute("SELECT niche, COUNT(*) as cnt FROM leads WHERE niche IS NOT NULL AND niche != '' GROUP BY niche").fetchall():
            niche_counts[row["niche"]] = row["cnt"]

        return {
            "total": total,
            "discovered": status_counts.get("discovered", 0),
            "qualified": status_counts.get("qualified", 0),
            "researched": status_counts.get("researched", 0),
            "enriched": status_counts.get("enriched", 0),
            "rejected": status_counts.get("rejected", 0),
            "by_canton": canton_counts,
            "by_niche": niche_counts,
        }
    finally:
        conn.close()


def export_leads_csv(filepath: str, include_all: bool = False) -> None:
    """Export leads to CSV. Defaults to the current delivered batch only
    (status='enriched') — not this run's rejected candidates (too small, no
    contact, duplicate, ranked below the cutoff) and never a past run's
    leftovers, since those no longer exist once a new run has wiped and
    replaced them (see start_new_run). Pass include_all=True to export
    literally everything still in the table, rejected rows included."""
    leads = get_leads() if include_all else get_leads(status="enriched")
    if not leads:
        logger.warning("No leads to export.")
        return

    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=leads[0].keys())
        writer.writeheader()
        writer.writerows(leads)

    logger.info(f"Exported {len(leads)} leads to {filepath}")


def export_leads_excel(filepath: str, include_all: bool = False) -> None:
    """Export leads to Excel. Same default as export_leads_csv: the current
    delivered batch only (status='enriched'), not rejected candidates or a
    past run's leftovers. Pass include_all=True for literally everything."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return

    leads = get_leads() if include_all else get_leads(status="enriched")
    if not leads:
        logger.warning("No leads to export.")
        return

    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header
    headers = list(leads[0].keys())
    ws.append(headers)

    # Style header
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")

    # Data
    for lead in leads:
        ws.append([lead.get(h, "") for h in headers])

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(filepath)
    logger.info(f"Exported {len(leads)} leads to {filepath}")


def clear_db() -> None:
    """Clear all records from the leads table."""
    conn = get_connection()
    try:
        conn.execute("DELETE FROM leads")
        conn.commit()
        logger.info("All leads cleared from database.")
    except Exception as e:
        logger.error(f"Error clearing database: {e}")
    finally:
        conn.close()


def start_new_run() -> str:
    """Called once at the very start of a pipeline run. Backs up the current
    database file, then wipes the `leads` table so every run starts from a
    clean, current-batch-only view instead of accumulating leads across runs
    forever — otherwise the dashboard fills up with stale leads from past runs
    with no way to tell which batch is current.

    Deliberately does NOT touch `blacklist`: that must persist across every
    wipe, forever, so a company already assessed in ANY past run is never
    rediscovered and re-delivered as if it were new — "old leads wipe out of
    view, but never come back and repeat" only holds if the memory of what's
    already been contacted survives the wipe.

    Returns a fresh run_id; the caller stamps it onto every lead this run
    touches so the current batch is traceable even if a future wipe is ever
    skipped (e.g. a --dashboard-only session)."""
    run_id = datetime.now().strftime("run_%Y%m%d_%H%M%S")

    try:
        if os.path.exists(_db_path):
            backup_dir = os.path.join(os.path.dirname(_db_path) or ".", "backups")
            os.makedirs(backup_dir, exist_ok=True)
            backup_path = os.path.join(backup_dir, f"leads_before_{run_id}.db")
            shutil.copy2(_db_path, backup_path)
            logger.info(f"Backed up database to {backup_path} before starting new run '{run_id}'")
    except Exception as e:
        logger.error(f"Could not back up database before new run (continuing anyway): {e}")

    conn = get_connection()
    try:
        conn.execute("DELETE FROM leads")
        conn.commit()
        logger.info(f"Cleared leads table for new run '{run_id}' (blacklist preserved).")
    except Exception as e:
        logger.error(f"Error clearing leads table for new run: {e}")
    finally:
        conn.close()

    return run_id


def get_current_run_info() -> Dict[str, Any]:
    """The run_id and timestamp of whatever batch is currently in the `leads`
    table, plus how many leads carry it — lets the dashboard show "this batch
    was generated at X" instead of leaving the user to guess which rows are
    current."""
    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT run_id, COUNT(*) as count, MAX(updated_at) as last_updated "
            "FROM leads WHERE run_id IS NOT NULL GROUP BY run_id ORDER BY last_updated DESC LIMIT 1"
        ).fetchone()
        if not row:
            return {"run_id": None, "count": 0, "last_updated": None}
        return {"run_id": row["run_id"], "count": row["count"], "last_updated": row["last_updated"]}
    finally:
        conn.close()
